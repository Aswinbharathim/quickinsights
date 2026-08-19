import asyncio
import io
import logging

import openpyxl
from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from app import db, rag, store
from app.config import DOCTYPE_FILTER
from app.models import (
    DatabaseConnection,
    DatabaseConnectionCreate,
    DatabaseConnectionOut,
    DatabaseConnectionUpdate,
    DbTable,
    DbTableUpdate,
    DescriptionImportProgress,
    DescriptionImportResponse,
    SchemaDiscoveryProgress,
    SchemaFieldDoc,
    TrainingJob,
    TrainSelectedRequest,
)
from app.schema_introspect import introspect_schema
from app.store import new_id_fn, now_fn
from app.vector_store import (
    collection_name_for_connection,
    delete_collection,
    ensure_collection,
    master_data_collection_name_for_connection,
    sql_examples_collection_name_for_connection,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/connections", tags=["connections"])

# Cap how many tables train concurrently. Without this, "Train all" on a large
# schema (hundreds of tables) fires that many simultaneous OpenAI + Qdrant
# calls at once, which reliably blows through OpenAI rate limits and used to
# race on creating a brand-new Qdrant collection (see vector_store.ensure_collection).
_TRAIN_CONCURRENCY = asyncio.Semaphore(5)


def _with_table_count(conn: DatabaseConnection) -> DatabaseConnectionOut:
    tables = store.tables_store.get(conn.id, {})
    trained_count = sum(1 for t in tables.values() if t.training_status == "trained")
    return DatabaseConnectionOut.from_connection(
        conn.model_copy(update={"table_count": len(tables), "trained_table_count": trained_count})
    )


@router.get("", response_model=list[DatabaseConnectionOut])
def list_connections():
    return [
        _with_table_count(c)
        for c in sorted(store.connections_store.values(), key=lambda c: c.created_at, reverse=True)
    ]


@router.post("", response_model=DatabaseConnectionOut, status_code=201)
def create_connection(payload: DatabaseConnectionCreate):
    new_id = new_id_fn()
    conn = DatabaseConnection(id=new_id, created_at=now_fn(), status="untested", **payload.model_dump())
    store.connections_store[new_id] = conn
    store.tables_store[new_id] = {}
    return _with_table_count(conn)


@router.put("/{conn_id}", response_model=DatabaseConnectionOut)
def update_connection(conn_id: str, payload: DatabaseConnectionUpdate):
    existing = store.connections_store.get(conn_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Connection not found")
    data = payload.model_dump()
    # The UI's password field is labeled "Leave blank to keep unchanged" when
    # editing — an empty string must NOT overwrite the stored password.
    if not data.get("password"):
        data.pop("password", None)
    updated = existing.model_copy(update={k: v for k, v in data.items() if v is not None})
    store.connections_store[conn_id] = updated
    return _with_table_count(updated)


@router.delete("/{conn_id}", status_code=204)
def delete_connection(conn_id: str):
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")

    # Every SQL example is now required to belong to a connection, and every
    # connection's examples live in their own Qdrant collection
    # (qi_sql_examples_<connection_id>) — so cleanup here is: capture
    # feedback-sourced examples first (need their source_message_id to clear
    # the chat message's feedback icon, which the DB cascade can't reach),
    # delete the connection row (the SqlExampleRow FK's ON DELETE CASCADE
    # takes care of the DB rows themselves), then drop that one Qdrant
    # collection wholesale instead of deleting each point individually.
    feedback_examples = [
        ex for ex in store.sql_example_store.values()
        if ex.connection_id == conn_id and ex.source_message_id
    ]

    del store.connections_store[conn_id]
    store.tables_store.pop(conn_id, None)
    store.training_job_store.delete_where(lambda j: j.connection_id == conn_id)
    store.discovery_progress_store.pop(conn_id, None)
    delete_collection(collection_name_for_connection(conn_id))
    delete_collection(sql_examples_collection_name_for_connection(conn_id))
    # MasterDataRecordRow.connection_id cascades the same way (ON DELETE
    # CASCADE) — this just drops its Qdrant collection the same way.
    delete_collection(master_data_collection_name_for_connection(conn_id))

    for ex in feedback_examples:
        try:
            store.clear_chat_message_feedback(ex.source_message_id)
        except Exception as e:
            print(f"Error clearing feedback on chat message {ex.source_message_id} -- {e}")
            logger.exception("Error clearing feedback on chat message %s", ex.source_message_id)


@router.post("/{conn_id}/test", response_model=DatabaseConnectionOut)
def test_connection(conn_id: str):
    """Step 1: verify connectivity only — does not touch the schema."""
    existing = store.connections_store.get(conn_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Connection not found")

    try:
        db.ping(existing)
    except Exception as e:
        print(f"Error testing connection {conn_id} -- {e}")
        logger.exception("Error testing connection %s", conn_id)
        store.connections_store[conn_id] = existing.model_copy(
            update={"status": "error", "last_tested_at": now_fn()}
        )
        raise HTTPException(status_code=400, detail=f"Could not connect: {e}")

    updated = existing.model_copy(update={"status": "connected", "last_tested_at": now_fn()})
    store.connections_store[conn_id] = updated
    return _with_table_count(updated)


def _merge_introspected_tables(conn_id: str, introspected: dict) -> None:
    existing_tables = store.tables_store.get(conn_id, {})
    tables: dict[str, DbTable] = {}
    for table_name, info in introspected.items():
        previous = existing_tables.get(table_name)
        new_columns = [
            SchemaFieldDoc(fieldname=c["fieldname"], label=c["fieldname"].replace("_", " ").title(), fieldtype=c["fieldtype"])
            for c in info["columns"]
        ]

        if previous:
            # Re-discovering an already-explored connection must not throw away
            # work already done on it (descriptions, training status/vectors) —
            # only structural facts (columns/FKs/row count) get refreshed.
            # Column descriptions the user wrote are preserved by fieldname.
            prev_descriptions = {c.fieldname: c.description for c in previous.columns if c.description}
            for col in new_columns:
                if col.fieldname in prev_descriptions:
                    col.description = prev_descriptions[col.fieldname]

            tables[table_name] = previous.model_copy(
                update={
                    "module": info.get("module") or previous.module,
                    "row_count": info["row_count"],
                    "is_child_table": info["is_child_table"],
                    "parent_table": info["parent_table"],
                    "foreign_keys": [fk for fk in info["foreign_keys"]],
                    "columns": new_columns,
                }
            )
        else:
            tables[table_name] = DbTable(
                table_name=table_name,
                doctype=info["doctype"],
                module=info.get("module"),
                description="",
                row_count=info["row_count"],
                primary_key="name",
                is_child_table=info["is_child_table"],
                parent_table=info["parent_table"],
                foreign_keys=[fk for fk in info["foreign_keys"]],
                columns=new_columns,
                training_status="untrained",
            )

    store.tables_store[conn_id] = tables


def _run_discovery_job(conn_id: str, conn: DatabaseConnection) -> None:
    """Runs on a worker thread (see discover_schema below) — updates
    discovery_progress_store as it goes so the frontend can show a real
    X/Y-tables progress bar instead of an indeterminate spinner."""
    progress = store.discovery_progress_store[conn_id]

    def on_progress(completed: int, total: int) -> None:
        store.discovery_progress_store[conn_id] = progress.model_copy(
            update={"completed": completed, "total": total}
        )

    try:
        raw_conn = db.get_raw_connection(conn)
        try:
            introspected = introspect_schema(
                raw_conn, conn.database_name, doctype_filter=DOCTYPE_FILTER or None, on_progress=on_progress
            )
        finally:
            raw_conn.close()

        _merge_introspected_tables(conn_id, introspected)

        store.discovery_progress_store[conn_id] = store.discovery_progress_store[conn_id].model_copy(
            update={
                "status": "completed",
                "completed_at": now_fn(),
                "table_count": len(introspected),
            }
        )
    except Exception as e:
        print(f"Error discovering schema for connection {conn_id} -- {e}")
        logger.exception("Error discovering schema for connection %s", conn_id)
        store.discovery_progress_store[conn_id] = store.discovery_progress_store[conn_id].model_copy(
            update={"status": "failed", "completed_at": now_fn(), "error": str(e)}
        )


@router.post("/{conn_id}/discover-schema", response_model=SchemaDiscoveryProgress, status_code=202)
async def discover_schema(conn_id: str):
    """Step 2: introspect the real database and populate its table list.

    Runs in the background and reports progress via the /discover-schema/progress
    endpoint below — real Frappe sites can have hundreds of tables, so this can
    take a few seconds and the UI polls for an X/Y-tables progress bar.
    """
    conn = store.connections_store.get(conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")

    progress = SchemaDiscoveryProgress(status="running", started_at=now_fn())
    store.discovery_progress_store[conn_id] = progress
    asyncio.create_task(asyncio.to_thread(_run_discovery_job, conn_id, conn))
    return progress


@router.get("/{conn_id}/discover-schema/progress", response_model=SchemaDiscoveryProgress)
def get_discovery_progress(conn_id: str):
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    return store.discovery_progress_store.get(conn_id, SchemaDiscoveryProgress())


@router.get("/{conn_id}/suggested-questions", response_model=list[str])
def get_suggested_questions(conn_id: str):
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    try:
        return rag.generate_suggested_questions(conn_id)
    except Exception as e:
        print(f"Error generating suggested questions for connection {conn_id} -- {e}")
        logger.exception("Error generating suggested questions for connection %s", conn_id)
        return []


@router.get("/{conn_id}/tables", response_model=list[DbTable])
def list_tables(conn_id: str, module: str | None = None):
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    tables = store.tables_store.get(conn_id, {}).values()
    if module:
        tables = [t for t in tables if t.module == module]
    return list(tables)


@router.get("/{conn_id}/tables/{table_name}", response_model=DbTable)
def get_table(conn_id: str, table_name: str):
    table = store.tables_store.get(conn_id, {}).get(table_name)
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    return table


@router.put("/{conn_id}/tables/{table_name}", response_model=DbTable)
def update_table(conn_id: str, table_name: str, payload: DbTableUpdate):
    tables = store.tables_store.get(conn_id)
    table = tables.get(table_name) if tables else None
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")

    update_dict = {}
    if payload.description is not None:
        update_dict["description"] = payload.description
    if payload.module is not None:
        update_dict["module"] = payload.module
    if payload.primary_key is not None:
        update_dict["primary_key"] = payload.primary_key
    if payload.is_child_table is not None:
        update_dict["is_child_table"] = payload.is_child_table
    if payload.parent_table is not None:
        update_dict["parent_table"] = payload.parent_table
    if payload.foreign_keys is not None:
        update_dict["foreign_keys"] = payload.foreign_keys
    if payload.columns is not None:
        update_dict["columns"] = payload.columns

    updated = table.model_copy(update=update_dict)

    # Column descriptions ("Edit schema descriptions" in ask-frappe.py) are
    # embedded as their own documentation vectors immediately, independent of
    # whether the table's full schema chunk has been trained yet — matches
    # ask-frappe.py's option 3 being usable on its own, before/without option 1.
    if payload.columns is not None:
        for col in updated.columns:
            if col.description:
                try:
                    rag.train_column_documentation_now(conn_id, table_name, col.fieldname, col.fieldtype, col.description)
                except Exception as e:
                    print(f"Error embedding column documentation for {table_name}.{col.fieldname} -- {e}")
                    logger.exception("Error embedding column documentation for %s.%s", table_name, col.fieldname)
                    # documentation embedding is supplementary — don't fail the save

    # If this table was already trained, also re-embed its schema chunk so
    # edited descriptions/relationships improve retrieval right away.
    if updated.training_status == "trained":
        tables[table_name] = updated
        try:
            vector_count = rag.train_table_now(conn_id, table_name)
            updated = updated.model_copy(
                update={"trained_at": now_fn(), "vector_count": vector_count, "training_error": None}
            )
        except Exception as e:
            print(f"Error re-embedding schema chunk for table {table_name} -- {e}")
            logger.exception("Error re-embedding schema chunk for table %s", table_name)
            updated = updated.model_copy(update={"training_status": "failed", "training_error": str(e)})

    tables[table_name] = updated
    return updated


async def _run_training_job(conn_id: str, table_name: str, job_id: str | None = None):
    tables = store.tables_store.get(conn_id)
    table = tables.get(table_name) if tables else None
    if not table:
        return
    tables[table_name] = table.model_copy(update={"training_status": "training"})

    succeeded = False
    try:
        async with _TRAIN_CONCURRENCY:
            vector_count = await asyncio.to_thread(rag.train_table_now, conn_id, table_name)
        table = tables[table_name]
        tables[table_name] = table.model_copy(
            update={
                "training_status": "trained",
                "trained_at": now_fn(),
                "vector_count": vector_count,
                "training_error": None,
            }
        )
        succeeded = True
    except Exception as e:
        print(f"Error training table {table_name} for connection {conn_id} -- {e}")
        logger.exception("Error training table %s for connection %s", table_name, conn_id)
        table = tables[table_name]
        tables[table_name] = table.model_copy(update={"training_status": "failed", "training_error": str(e)})

    if job_id:
        job = store.training_job_store.get(job_id)
        if job:
            completed = job.completed + 1
            failed = job.failed if succeeded else job.failed + 1
            update = {"completed": completed, "failed": failed}
            if completed >= job.total:
                update.update({"status": "completed", "completed_at": now_fn()})
            store.training_job_store[job_id] = job.model_copy(update=update)


@router.post("/{conn_id}/tables/{table_name}/train", response_model=DbTable, status_code=202)
async def train_table(conn_id: str, table_name: str):
    tables = store.tables_store.get(conn_id)
    table = tables.get(table_name) if tables else None
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    updated = table.model_copy(update={"training_status": "queued", "training_error": None})
    tables[table_name] = updated

    # Create the collection up front, outside the concurrent fan-out below —
    # see ensure_collection's docstring on why concurrent callers racing to
    # create the same brand-new collection used to fail training.
    try:
        await asyncio.to_thread(ensure_collection, collection_name_for_connection(conn_id))
    except Exception as e:
        print(f"Error creating vector store collection for connection {conn_id} -- {e}")
        logger.exception("Error creating vector store collection for connection %s", conn_id)
        tables[table_name] = updated.model_copy(update={"training_status": "failed", "training_error": str(e)})
        raise HTTPException(status_code=502, detail=f"Vector store unreachable: {e}")

    job_id = new_id_fn()
    store.training_job_store[job_id] = TrainingJob(
        id=job_id,
        connection_id=conn_id,
        table_names=[table_name],
        total=1,
        completed=0,
        status="running",
        started_at=now_fn(),
    )
    # asyncio.create_task (not FastAPI's BackgroundTasks, which awaits tasks one at a
    # time) so many tables train concurrently instead of queueing up sequentially.
    asyncio.create_task(_run_training_job(conn_id, table_name, job_id))
    return updated


async def _queue_training_job(conn_id: str, tables: dict[str, DbTable], to_train: list[str]) -> TrainingJob:
    """Shared by train-all and train-selected: creates the collection up front
    (avoids the concurrent-create race — see ensure_collection), records a
    TrainingJob, and fans out one background task per table."""
    if to_train:
        try:
            await asyncio.to_thread(ensure_collection, collection_name_for_connection(conn_id))
        except Exception as e:
            print(f"Error creating vector store collection for connection {conn_id} -- {e}")
            logger.exception("Error creating vector store collection for connection %s", conn_id)
            raise HTTPException(status_code=502, detail=f"Vector store unreachable: {e}")

    job_id = new_id_fn()
    job = TrainingJob(
        id=job_id,
        connection_id=conn_id,
        table_names=to_train,
        total=len(to_train),
        completed=0,
        status="completed" if not to_train else "running",
        started_at=now_fn(),
        completed_at=now_fn() if not to_train else None,
    )
    store.training_job_store[job_id] = job

    for name in to_train:
        tables[name] = tables[name].model_copy(update={"training_status": "queued", "training_error": None})
        asyncio.create_task(_run_training_job(conn_id, name, job_id))

    return job


@router.post("/{conn_id}/train-all", response_model=TrainingJob, status_code=202)
async def train_all_tables(conn_id: str, force: bool = False):
    """Step 3: queue background training for tables that need it, tracked as a job.

    force=True re-trains EVERY table regardless of current status — used by the
    "Retrain" action once a connection is already fully (or partially) trained.
    """
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    tables = store.tables_store.get(conn_id)
    if not tables:
        raise HTTPException(
            status_code=400,
            detail="No schema discovered yet — discover the schema first.",
        )

    if force:
        to_train = [name for name, t in tables.items() if t.training_status not in ("queued", "training")]
    else:
        to_train = [
            name for name, t in tables.items() if t.training_status not in ("trained", "queued", "training")
        ]

    return await _queue_training_job(conn_id, tables, to_train)


@router.post("/{conn_id}/train-selected", response_model=TrainingJob, status_code=202)
async def train_selected_tables(conn_id: str, payload: TrainSelectedRequest):
    """Train exactly the given tables (checked in the Schema Explorer list) —
    the UI alternative to the .env DOCTYPE_FILTER: discovery still pulls every
    table, but you choose which ones actually get embedded, so you don't burn
    tokens training tables you don't need for testing.

    Unlike train-all, an explicit selection always (re)trains the picked
    tables regardless of current status — selecting an already-trained table
    re-embeds it.
    """
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    tables = store.tables_store.get(conn_id)
    if not tables:
        raise HTTPException(
            status_code=400,
            detail="No schema discovered yet — discover the schema first.",
        )

    unknown = [name for name in payload.table_names if name not in tables]
    if unknown:
        raise HTTPException(status_code=404, detail=f"Table(s) not found: {', '.join(unknown)}")

    to_train = [
        name for name in payload.table_names if tables[name].training_status not in ("queued", "training")
    ]

    return await _queue_training_job(conn_id, tables, to_train)


@router.get("/{conn_id}/training-jobs", response_model=list[TrainingJob])
def list_training_jobs(conn_id: str):
    if conn_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    jobs = [j for j in store.training_job_store.values() if j.connection_id == conn_id]
    return sorted(jobs, key=lambda j: j.started_at, reverse=True)


@router.get("/{conn_id}/export-descriptions")
def export_descriptions(conn_id: str, module: str | None = None):
    conn = store.connections_store.get(conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")

    tables_dict = store.tables_store.get(conn_id, {})
    if not tables_dict:
        raise HTTPException(status_code=400, detail="No schema tables found for this connection yet.")

    # Filter by module if specified (and not "all")
    if module and module.strip() and module.strip().lower() != "all":
        mod_filter = module.strip().lower()
        target_tables = {k: v for k, v in tables_dict.items() if (v.module or "").strip().lower() == mod_filter}
    else:
        target_tables = tables_dict

    wb = openpyxl.Workbook()

    # Sheet 1: Table_Descriptions
    ws_table = wb.active
    ws_table.title = "Table_Descriptions"
    ws_table.append(["Table Name", "DocType", "Module", "Description"])

    # Sheet 2: Column_Descriptions
    ws_col = wb.create_sheet(title="Column_Descriptions")
    ws_col.append(["Table Name", "Field Name", "Field Type", "Description"])

    for table_name in sorted(target_tables.keys()):
        table = target_tables[table_name]
        ws_table.append([
            table.table_name,
            table.doctype or "",
            table.module or "",
            table.description or "",
        ])
        for col in table.columns:
            ws_col.append([
                table.table_name,
                col.fieldname,
                col.fieldtype or "",
                col.description or "",
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"table_descriptions_{conn.database_name}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _run_import_job(conn_id: str, file_bytes: bytes):
    progress = store.import_progress_store.get(conn_id)
    if not progress:
        return

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        logger.exception("Failed to parse Excel file for connection %s", conn_id)
        progress.status = "failed"
        progress.error = f"Failed to parse Excel file: {str(e)}"
        progress.completed_at = now_fn()
        return

    tables_dict = store.tables_store.get(conn_id)
    if not tables_dict:
        progress.status = "failed"
        progress.error = "No tables discovered for this connection yet."
        progress.completed_at = now_fn()
        return

    sheet_map = {name.lower().strip(): name for name in wb.sheetnames}
    table_sheet_name = sheet_map.get("table_descriptions") or sheet_map.get("tables") or wb.sheetnames[0]
    col_sheet_name = sheet_map.get("column_descriptions") or sheet_map.get("columns") or (wb.sheetnames[1] if len(wb.sheetnames) > 1 else None)

    table_rows_data: list[dict[str, str]] = []
    if table_sheet_name in wb.sheetnames:
        sheet = wb[table_sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            headers = [str(cell or "").strip().lower() for cell in rows[0]]
            for r_idx, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue
                row_dict: dict[str, str] = {}
                for c_idx, cell in enumerate(row):
                    if c_idx < len(headers):
                        row_dict[headers[c_idx]] = str(cell or "").strip() if cell is not None else ""
                row_dict["_row_num"] = str(r_idx)
                table_rows_data.append(row_dict)

    col_rows_data: list[dict[str, str]] = []
    if col_sheet_name and col_sheet_name in wb.sheetnames:
        sheet = wb[col_sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            headers = [str(cell or "").strip().lower() for cell in rows[0]]
            for r_idx, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue
                row_dict = {}
                for c_idx, cell in enumerate(row):
                    if c_idx < len(headers):
                        row_dict[headers[c_idx]] = str(cell or "").strip() if cell is not None else ""
                row_dict["_row_num"] = str(r_idx)
                col_rows_data.append(row_dict)

    total_rows = len(table_rows_data) + len(col_rows_data)
    progress.total_rows = total_rows
    progress.processed_rows = 0

    errors: list[str] = []
    updated_tables: set[str] = set()
    updated_columns_count = 0
    tables_to_retrain: set[str] = set()
    conn_obj = store.connections_store.get(conn_id)

    # Process Table_Descriptions
    for row in table_rows_data:
        r_num = row.get("_row_num", "?")
        table_name = row.get("table name") or row.get("table_name") or row.get("tablename") or row.get("table") or ""
        desc = row.get("description") or row.get("table description") or ""

        if not table_name:
            errors.append(f"Table_Descriptions Row {r_num}: Missing 'Table Name'")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        if table_name not in tables_dict:
            errors.append(f"Table_Descriptions Row {r_num}: Table '{table_name}' does not exist in connection '{conn_obj.database_name if conn_obj else conn_id}'")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        table = tables_dict[table_name]
        curr_desc = table.description or ""
        if desc != curr_desc:
            tables_dict[table_name] = table.model_copy(update={"description": desc})
            updated_tables.add(table_name)
            tables_to_retrain.add(table_name)

        progress.processed_rows += 1

    # Process Column_Descriptions
    for row in col_rows_data:
        r_num = row.get("_row_num", "?")
        table_name = row.get("table name") or row.get("table_name") or row.get("tablename") or row.get("table") or ""
        field_name = row.get("field name") or row.get("field_name") or row.get("fieldname") or row.get("column") or row.get("column name") or ""
        desc = row.get("description") or row.get("field description") or row.get("column description") or ""

        if not table_name:
            errors.append(f"Column_Descriptions Row {r_num}: Missing 'Table Name'")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        if table_name not in tables_dict:
            errors.append(f"Column_Descriptions Row {r_num}: Table '{table_name}' does not exist in connection '{conn_obj.database_name if conn_obj else conn_id}'")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        table = tables_dict[table_name]
        col_match = next((c for c in table.columns if c.fieldname == field_name), None)
        if not col_match:
            errors.append(f"Column_Descriptions Row {r_num}: Field '{field_name}' does not exist in table '{table_name}'")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        curr_col_desc = col_match.description or ""
        if desc != curr_col_desc:
            new_columns = [
                c.model_copy(update={"description": desc}) if c.fieldname == field_name else c
                for c in table.columns
            ]
            table_updated = table.model_copy(update={"columns": new_columns})
            tables_dict[table_name] = table_updated
            updated_columns_count += 1
            updated_tables.add(table_name)
            tables_to_retrain.add(table_name)

            if desc:
                try:
                    rag.train_column_documentation_now(
                        conn_id, table_name, field_name, col_match.fieldtype or "Data", desc, module=table.module
                    )
                except Exception as e:
                    logger.exception("Error embedding column doc %s.%s", table_name, field_name)

        progress.processed_rows += 1

    # Retrain/re-embed schema chunks for affected trained tables
    retrained_tables_count = 0
    for table_name in sorted(tables_to_retrain):
        table = tables_dict[table_name]
        if table.training_status == "trained":
            try:
                vector_count = rag.train_table_now(conn_id, table_name)
                tables_dict[table_name] = table.model_copy(
                    update={"trained_at": now_fn(), "vector_count": vector_count, "training_error": None}
                )
                retrained_tables_count += 1
            except Exception as e:
                logger.exception("Error re-embedding table %s", table_name)
                tables_dict[table_name] = table.model_copy(
                    update={"training_status": "failed", "training_error": str(e)}
                )

    progress.status = "completed"
    progress.updated_tables_count = len(updated_tables)
    progress.updated_columns_count = updated_columns_count
    progress.retrained_tables_count = retrained_tables_count
    progress.errors = errors
    progress.completed_at = now_fn()


@router.post("/{conn_id}/import-descriptions", response_model=DescriptionImportProgress, status_code=202)
async def import_descriptions(conn_id: str, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    conn = store.connections_store.get(conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")

    tables_dict = store.tables_store.get(conn_id)
    if not tables_dict:
        raise HTTPException(status_code=400, detail="No tables discovered for this connection yet.")

    filename = file.filename or ""
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel (.xlsx) file.")

    content = await file.read()

    progress = DescriptionImportProgress(
        status="running",
        started_at=now_fn(),
    )
    store.import_progress_store[conn_id] = progress

    background_tasks.add_task(_run_import_job, conn_id, content)

    return progress


@router.get("/{conn_id}/import-descriptions/progress", response_model=DescriptionImportProgress)
def get_import_descriptions_progress(conn_id: str):
    conn = store.connections_store.get(conn_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Connection not found")
    return store.import_progress_store.get(conn_id, DescriptionImportProgress())


