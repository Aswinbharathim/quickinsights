import csv
import io
import logging
from typing import Optional

import openpyxl
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app import rag, store
from app.models import (
    SqlExample,
    SqlExampleCreate,
    SqlExampleUpdate,
    SqlExampleImportProgress,
    SqlExampleBatchDeleteRequest,
)
from app.store import new_id_fn, now_fn

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/sql-examples", tags=["sql-examples"])


@router.get("", response_model=list[SqlExample])
def list_sql_examples(module: Optional[str] = None, connection_id: Optional[str] = None):
    """Only confirmed examples — rejected (thumbs-down) ones are kept as a
    retrieval-time "don't repeat this" signal but aren't manually manageable
    training examples, so they're excluded from this list.

    connection_id filters to exactly that connection's examples — no more
    "global" examples exist, so there's no fallback merge here; omitting
    connection_id (the "All Connections" view) simply shows every
    connection's examples together."""
    examples = (e for e in store.sql_example_store.values() if e.vote == "up")
    if module:
        examples = (e for e in examples if e.module == module)
    if connection_id:
        examples = (e for e in examples if e.connection_id == connection_id)
    return sorted(
        examples,
        key=lambda e: e.created_at,
        reverse=True,
    )


_TEMPLATE_ROWS = [
    (
        "Show me top 10 patient appointments today",
        "SELECT name, patient, appointment_time FROM `tabPatient Appointment` "
        "WHERE appointment_date = CURDATE() ORDER BY appointment_time ASC LIMIT 10;",
        "Healthcare",
        "patient, appointment, healthcare",
    ),
    (
        "Count total active orders by status",
        "SELECT status, COUNT(*) AS order_count FROM `tabSales Order` "
        "WHERE status != 'Cancelled' GROUP BY status;",
        "Selling",
        "orders, summary, selling",
    ),
]


def _build_sql_examples_workbook(rows: list[tuple[str, str, str, str]]) -> io.BytesIO:
    """Single source of truth for the .xlsx shape both /export and /template
    produce — exactly the columns /import parses (Question, SQL, Module,
    Tags), so a file downloaded from either endpoint round-trips back
    through /import without any column-mapping surprises."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SQL Examples"
    ws.append(["Question", "SQL", "Module", "Tags"])
    for row in rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/template")
def download_sql_examples_template():
    """Always the same 2-row starter template, regardless of connection —
    what "Export"/"Download Template" both call, on the SQL Examples page,
    the Connections page pipeline's import dialog, and anywhere else that
    needs the expected columns to fill in and re-upload via /import. Not
    connection-scoped and never the actual saved examples — deliberately
    template-only, so there's one canonical definition of the format
    instead of a separate hardcoded copy per caller."""
    buffer = _build_sql_examples_workbook(_TEMPLATE_ROWS)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sql_examples_template.xlsx"},
    )


@router.post("", response_model=SqlExample, status_code=201)
def create_sql_example(payload: SqlExampleCreate):
    if payload.connection_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    new_id = new_id_fn()
    record = SqlExample(id=new_id, created_at=now_fn(), source="manual", **payload.model_dump())
    store.sql_example_store[new_id] = record
    try:
        rag.train_sql_example_now(
            new_id, record.question, record.sql, module=record.module, connection_id=record.connection_id
        )
    except Exception as e:
        print(f"Error embedding SQL example {new_id} -- {e}")
        logger.exception("Error embedding SQL example %s", new_id)
    return record


@router.post("/batch", response_model=list[SqlExample], status_code=201)
def create_sql_examples_batch(payload: list[SqlExampleCreate]):
    for item in payload:
        if item.connection_id not in store.connections_store:
            raise HTTPException(status_code=404, detail=f"Connection not found: {item.connection_id}")
    created_records = []
    for item in payload:
        new_id = new_id_fn()
        record = SqlExample(id=new_id, created_at=now_fn(), source="excel_import", **item.model_dump())
        store.sql_example_store[new_id] = record
        try:
            rag.train_sql_example_now(
                new_id, record.question, record.sql, module=record.module, connection_id=record.connection_id
            )
        except Exception as e:
            print(f"Error embedding SQL example {new_id} -- {e}")
            logger.exception("Error embedding SQL example %s", new_id)
        created_records.append(record)
    return created_records


def _parse_sql_example_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    rows_data: list[dict[str, str]] = []
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")
        headers = [str(cell or "").strip().lower() for cell in rows[0]]
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = str(cell or "").strip()
            rows_data.append(row_dict)
    else:
        # Assume CSV
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            cleaned_row = {str(k or "").strip().lower(): str(v or "").strip() for k, v in row.items()}
            if any(cleaned_row.values()):
                rows_data.append(cleaned_row)
    return rows_data


def _run_sql_example_import_job(connection_id: str, filename: str, content: bytes) -> None:
    """Runs via BackgroundTasks (see import_sql_examples_file below) — after
    the request that queued it has already returned, so a big spreadsheet's
    parsing + per-row embedding calls never block the HTTP response or the
    page that started the import. Mirrors connections.py's
    _run_import_job (bulk table/column-description import) exactly, one
    level simpler since there's only one sheet/row shape here."""
    progress = store.sql_example_import_progress_store.get(connection_id)
    if not progress:
        return

    try:
        rows_data = _parse_sql_example_rows(filename, content)
    except Exception as e:
        print(f"Error parsing SQL example import file for connection {connection_id} -- {e}")
        logger.exception("Error parsing SQL example import file for connection %s", connection_id)
        progress.status = "failed"
        progress.error = f"Failed to parse file: {e}"
        progress.completed_at = now_fn()
        return

    if not rows_data:
        progress.status = "failed"
        progress.error = "No data rows found in file"
        progress.completed_at = now_fn()
        return

    progress.total_rows = len(rows_data)
    errors: list[str] = []
    imported_count = 0

    for i, row in enumerate(rows_data, start=2):
        # Header alias resolution
        question = row.get("question") or row.get("queries") or row.get("query") or row.get("prompt") or ""
        sql = row.get("sql") or row.get("sql query") or row.get("sql_query") or row.get("code") or ""
        module_val = row.get("module") or row.get("frappe_module") or None
        raw_tags = row.get("tags") or row.get("tag") or row.get("category") or ""

        if not question or not sql:
            errors.append(f"Row {i}: Missing required fields ('question' and 'sql')")
            progress.failed_count += 1
            progress.processed_rows += 1
            continue

        tags = [t.strip() for t in raw_tags.replace(";", ",").split(",") if t.strip()]

        new_id = new_id_fn()
        record = SqlExample(
            id=new_id,
            question=question,
            sql=sql,
            module=module_val,
            connection_id=connection_id,
            tags=tags,
            created_at=now_fn(),
            source="excel_import",
            vote="up",
        )
        store.sql_example_store[new_id] = record
        try:
            rag.train_sql_example_now(
                new_id, record.question, record.sql, module=record.module, connection_id=record.connection_id
            )
        except Exception as e:
            print(f"Error embedding SQL example {new_id} -- {e}")
            logger.exception("Error embedding SQL example %s", new_id)
        imported_count += 1
        progress.processed_rows += 1

    progress.status = "completed"
    progress.imported_count = imported_count
    progress.errors = errors
    progress.completed_at = now_fn()


@router.post("/import", response_model=SqlExampleImportProgress, status_code=202)
async def import_sql_examples_file(
    background_tasks: BackgroundTasks, file: UploadFile = File(...), connection_id: str = Form(...)
):
    """connection_id is required, not optional — every imported example must
    be tied to the connection it was imported for. There is no "global"
    concept anywhere anymore, for either manual entries or bulk import.

    Returns immediately (202) with a "running" SqlExampleImportProgress —
    the actual parsing + per-row embedding happens in the background
    (_run_sql_example_import_job); poll GET /import/progress?connection_id=
    for status. Only one import can be in flight per connection at a time —
    starting another overwrites the previous run's progress record."""
    if connection_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")

    filename = file.filename or ""
    content = await file.read()

    progress = SqlExampleImportProgress(status="running", started_at=now_fn())
    store.sql_example_import_progress_store[connection_id] = progress
    background_tasks.add_task(_run_sql_example_import_job, connection_id, filename, content)
    return progress


@router.get("/import/progress", response_model=SqlExampleImportProgress)
def get_sql_example_import_progress(connection_id: str):
    if connection_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    return store.sql_example_import_progress_store.get(connection_id, SqlExampleImportProgress())


@router.put("/{item_id}", response_model=SqlExample)
def update_sql_example(item_id: str, payload: SqlExampleUpdate):
    existing = store.sql_example_store.get(item_id)
    if not existing:
        raise HTTPException(status_code=404, detail="SQL example not found")
    if payload.connection_id is not None and payload.connection_id not in store.connections_store:
        raise HTTPException(status_code=404, detail="Connection not found")
    updated = existing.model_copy(update={k: v for k, v in payload.model_dump().items() if v is not None})
    store.sql_example_store[item_id] = updated
    # Each connection has its own Qdrant collection now, so re-training after
    # a connection_id change must also drop the stale point left behind in
    # the OLD connection's collection — upserting into the new one alone
    # would leave a duplicate/orphaned vector there forever.
    if updated.connection_id != existing.connection_id:
        try:
            rag.delete_sql_example_vector(item_id, existing.connection_id)
        except Exception as e:
            print(f"Error deleting stale SQL example vector {item_id} -- {e}")
            logger.exception("Error deleting stale SQL example vector %s", item_id)
    try:
        rag.train_sql_example_now(
            item_id, updated.question, updated.sql, module=updated.module, connection_id=updated.connection_id
        )
    except Exception as e:
        print(f"Error re-embedding SQL example {item_id} -- {e}")
        logger.exception("Error re-embedding SQL example %s", item_id)
    return updated


@router.delete("/{item_id}", status_code=204)
def delete_sql_example(item_id: str):
    existing = store.sql_example_store.get(item_id)
    if not existing:
        raise HTTPException(status_code=404, detail="SQL example not found")
    del store.sql_example_store[item_id]
    try:
        rag.delete_sql_example_vector(item_id, existing.connection_id)
    except Exception as e:
        print(f"Error deleting SQL example {item_id} from Qdrant -- {e}")
        logger.exception("Error deleting SQL example %s from Qdrant", item_id)
    if existing.source_message_id:
        try:
            store.clear_chat_message_feedback(existing.source_message_id)
        except Exception as e:
            print(f"Error clearing feedback on chat message {existing.source_message_id} -- {e}")
            logger.exception("Error clearing feedback on chat message %s", existing.source_message_id)


@router.post("/batch-delete", status_code=204)
def delete_sql_examples_batch(payload: SqlExampleBatchDeleteRequest):
    for item_id in payload.ids:
        existing = store.sql_example_store.get(item_id)
        if existing:
            del store.sql_example_store[item_id]
            try:
                rag.delete_sql_example_vector(item_id, existing.connection_id)
            except Exception as e:
                print(f"Error deleting SQL example {item_id} from Qdrant -- {e}")
                logger.exception("Error deleting SQL example %s from Qdrant", item_id)
            if existing.source_message_id:
                try:
                    store.clear_chat_message_feedback(existing.source_message_id)
                except Exception as e:
                    print(f"Error clearing feedback on chat message {existing.source_message_id} -- {e}")
                    logger.exception("Error clearing feedback on chat message %s", existing.source_message_id)

